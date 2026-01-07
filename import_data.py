#!/usr/bin/env python
"""
Script para importar datos desde archivos Excel a la base de datos.
Uso: python import_data.py [--clear]

El argumento --clear elimina los datos existentes antes de importar.
"""
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from app import create_app
from models import db, Product, ProductType, Customer, Location, Lot, LotLocation


def parse_date(value):
    """Parse date from Excel or string format"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def parse_float(value):
    """Parse float, handling None and strings"""
    if value is None or value == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_product_type(value):
    """Parse product type string to enum"""
    if value is None:
        return None
    value = str(value).lower().strip()
    mapping = {
        'raw_material': ProductType.RAW_MATERIAL,
        'materia_prima': ProductType.RAW_MATERIAL,
        'mp': ProductType.RAW_MATERIAL,
        'packaging': ProductType.PACKAGING,
        'envase': ProductType.PACKAGING,
        'env': ProductType.PACKAGING,
        'finished_product': ProductType.FINISHED_PRODUCT,
        'producto_acabado': ProductType.FINISHED_PRODUCT,
        'pa': ProductType.FINISHED_PRODUCT,
    }
    return mapping.get(value)


def import_ubicaciones(filepath):
    """Import locations from Excel"""
    if not os.path.exists(filepath):
        print(f"  ⚠ Archivo no encontrado: {filepath}")
        return 0
    
    wb = load_workbook(filepath)
    ws = wb.active
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        code, name, is_available = row[0], row[1], row[2] if len(row) > 2 else False
        
        # Check if exists
        existing = Location.query.filter_by(code=code).first()
        if existing:
            print(f"    - Ubicación {code} ya existe, actualizando...")
            existing.name = name
            existing.is_available = bool(is_available)
        else:
            location = Location(
                code=str(code),
                name=str(name),
                is_available=bool(is_available) if is_available else False
            )
            db.session.add(location)
            count += 1
    
    db.session.commit()
    print(f"  ✓ {count} ubicaciones importadas")
    return count


def import_productos(filepath):
    """Import products from Excel"""
    if not os.path.exists(filepath):
        print(f"  ⚠ Archivo no encontrado: {filepath}")
        return 0
    
    wb = load_workbook(filepath)
    ws = wb.active
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        code = str(row[0])
        name = str(row[1]) if row[1] else ''
        product_type = parse_product_type(row[2])
        description = str(row[3]) if row[3] else ''
        min_stock = parse_float(row[4])
        storage_unit = str(row[5]) if row[5] else 'ud'
        consumption_unit = str(row[6]) if len(row) > 6 and row[6] else None
        density = parse_float(row[7]) if len(row) > 7 else None
        
        if not product_type:
            print(f"    ⚠ Tipo de producto inválido para {code}, saltando...")
            continue
        
        # Check if exists
        existing = Product.query.filter_by(code=code).first()
        if existing:
            print(f"    - Producto {code} ya existe, actualizando...")
            existing.name = name
            existing.type = product_type
            existing.description = description
            existing.min_stock = min_stock
            existing.storage_unit = storage_unit
            existing.consumption_unit = consumption_unit
            existing.density = density
        else:
            product = Product(
                code=code,
                name=name,
                type=product_type,
                description=description,
                min_stock=min_stock,
                storage_unit=storage_unit,
                consumption_unit=consumption_unit,
                density=density
            )
            db.session.add(product)
            count += 1
    
    db.session.commit()
    print(f"  ✓ {count} productos importados")
    return count


def import_clientes(filepath):
    """Import customers from Excel"""
    if not os.path.exists(filepath):
        print(f"  ⚠ Archivo no encontrado: {filepath}")
        return 0
    
    wb = load_workbook(filepath)
    ws = wb.active
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        code = str(row[0])
        name = str(row[1]) if row[1] else ''
        email = str(row[2]) if len(row) > 2 and row[2] else None
        phone = str(row[3]) if len(row) > 3 and row[3] else None
        address = str(row[4]) if len(row) > 4 and row[4] else None
        
        # Check if exists
        existing = Customer.query.filter_by(code=code).first()
        if existing:
            print(f"    - Cliente {code} ya existe, actualizando...")
            existing.name = name
            existing.email = email
            existing.phone = phone
            existing.address = address
        else:
            customer = Customer(
                code=code,
                name=name,
                email=email,
                phone=phone,
                address=address
            )
            db.session.add(customer)
            count += 1
    
    db.session.commit()
    print(f"  ✓ {count} clientes importados")
    return count


def import_inventario(filepath):
    """Import inventory (lots) from Excel"""
    if not os.path.exists(filepath):
        print(f"  ⚠ Archivo no encontrado: {filepath}")
        return 0
    
    wb = load_workbook(filepath)
    ws = wb.active
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        product_code = str(row[0])
        lot_number = str(row[1])
        quantity = parse_float(row[2])
        unit = str(row[3]) if row[3] else 'ud'
        manufacturing_date = parse_date(row[4])
        expiration_date = parse_date(row[5]) if len(row) > 5 else None
        location_code = str(row[6]) if len(row) > 6 and row[6] else 'LIB'
        
        if quantity is None:
            print(f"    ⚠ Cantidad inválida para lote {lot_number}, saltando...")
            continue
        
        if manufacturing_date is None:
            manufacturing_date = datetime.now().date()
        
        # Find product
        product = Product.query.filter_by(code=product_code).first()
        if not product:
            print(f"    ⚠ Producto {product_code} no encontrado, saltando lote {lot_number}...")
            continue
        
        # Find location
        location = Location.query.filter_by(code=location_code).first()
        if not location:
            print(f"    ⚠ Ubicación {location_code} no encontrada, usando LIB...")
            location = Location.query.filter_by(code='LIB').first()
            if not location:
                print(f"    ⚠ Ubicación LIB no existe, saltando lote {lot_number}...")
                continue
        
        # Check if lot exists
        existing = Lot.query.filter_by(product_id=product.id, lot_number=lot_number).first()
        if existing:
            print(f"    - Lote {lot_number} ya existe para {product_code}, saltando...")
            continue
        
        # Create lot
        lot = Lot(
            product_id=product.id,
            lot_number=lot_number,
            manufacturing_date=manufacturing_date,
            expiration_date=expiration_date,
            initial_quantity=quantity,
            current_quantity=quantity,
            unit=unit
        )
        db.session.add(lot)
        db.session.flush()
        
        # Create lot location
        lot_location = LotLocation(
            lot_id=lot.id,
            location_id=location.id,
            quantity=quantity
        )
        db.session.add(lot_location)
        count += 1
    
    db.session.commit()
    print(f"  ✓ {count} lotes importados")
    return count


def clear_data():
    """Clear all data from the database"""
    print("⚠ Eliminando datos existentes...")
    LotLocation.query.delete()
    Lot.query.delete()
    Customer.query.delete()
    Product.query.delete()
    Location.query.delete()
    db.session.commit()
    print("  ✓ Datos eliminados")


def main():
    """Main import function"""
    clear = '--clear' in sys.argv
    
    app = create_app('development')
    
    with app.app_context():
        print("=" * 50)
        print("IMPORTACIÓN DE DATOS DESDE EXCEL")
        print("=" * 50)
        
        if clear:
            clear_data()
        
        base_path = 'plantillas_importacion'
        
        print("\n1. Importando ubicaciones...")
        import_ubicaciones(os.path.join(base_path, 'ubicaciones.xlsx'))
        
        print("\n2. Importando productos...")
        import_productos(os.path.join(base_path, 'productos.xlsx'))
        
        print("\n3. Importando clientes...")
        import_clientes(os.path.join(base_path, 'clientes.xlsx'))
        
        print("\n4. Importando inventario (lotes)...")
        import_inventario(os.path.join(base_path, 'inventario.xlsx'))
        
        print("\n" + "=" * 50)
        print("IMPORTACIÓN COMPLETADA")
        print("=" * 50)
        print(f"  - Ubicaciones: {Location.query.count()}")
        print(f"  - Productos: {Product.query.count()}")
        print(f"  - Clientes: {Customer.query.count()}")
        print(f"  - Lotes: {Lot.query.count()}")


if __name__ == '__main__':
    main()
