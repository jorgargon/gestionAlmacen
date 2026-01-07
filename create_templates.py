#!/usr/bin/env python
"""
Script para generar plantillas Excel vacías con ejemplos.
Uso: python create_templates.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


def style_header(ws, num_cols):
    """Apply header styling"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def auto_width(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_ubicaciones():
    """Create locations template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ubicaciones"
    
    # Headers
    headers = ["Código", "Nombre", "Disponible para uso"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Example data
    examples = [
        ("REC", "Recepción - Cuarentena", False),
        ("LIB", "Liberado - Disponible", True),
        ("DEV", "Devoluciones", False),
        ("NC", "No Conforme", False),
        ("FAB", "Fabricación", False),
    ]
    
    for row, data in enumerate(examples, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    style_header(ws, len(headers))
    auto_width(ws)
    
    wb.save("plantillas_importacion/ubicaciones.xlsx")
    print("  ✓ ubicaciones.xlsx creado")


def create_productos():
    """Create products template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Headers
    headers = ["Código", "Nombre", "Tipo", "Descripción", "Stock Mínimo", 
               "Unidad Almacén", "Unidad Consumo", "Densidad (kg/l)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Instructions row
    instructions = [
        "Ej: MP-001",
        "Nombre completo",
        "MP/ENV/PA",
        "Opcional",
        "Opcional",
        "kg/l/ud",
        "kg/g/ud (opcional)",
        "Solo para líquidos"
    ]
    for col, inst in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col, value=inst)
        cell.font = Font(italic=True, color="808080")
    
    # Example data
    examples = [
        ("MP-001", "Agua destilada", "MP", "Agua destilada para uso cosmético", 100, "l", "kg", 1.0),
        ("MP-002", "Glicerina vegetal", "MP", "Glicerina 99.5%", 50, "kg", "kg", None),
        ("MP-003", "Aceite esencial lavanda", "MP", "Aceite 100% puro", 5, "l", "kg", 0.985),
        ("ENV-001", "Tarro 50ml", "ENV", "Tarro de vidrio", 500, "ud", "ud", None),
        ("ENV-002", "Etiqueta adhesiva", "ENV", "Etiqueta para tarros", 500, "ud", "ud", None),
        ("PA-001", "Crema Hidratante 50ml", "PA", "Crema facial hidratante", 50, "ud", "ud", None),
    ]
    
    for row, data in enumerate(examples, 3):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    style_header(ws, len(headers))
    auto_width(ws)
    
    wb.save("plantillas_importacion/productos.xlsx")
    print("  ✓ productos.xlsx creado")


def create_clientes():
    """Create customers template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = ["Código", "Nombre", "Email", "Teléfono", "Dirección"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Example data
    examples = [
        ("CLI-001", "Farmacia Central", "farmacia@example.com", "+34 900 123 456", "Calle Principal 123, Barcelona"),
        ("CLI-002", "Perfumería Elegance", "info@elegance.com", "+34 900 123 457", "Av. Diagonal 456, Barcelona"),
        ("CLI-003", "Spa Wellness", "contacto@spawellness.com", "+34 900 123 458", "Paseo de Gracia 789, Barcelona"),
    ]
    
    for row, data in enumerate(examples, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    style_header(ws, len(headers))
    auto_width(ws)
    
    wb.save("plantillas_importacion/clientes.xlsx")
    print("  ✓ clientes.xlsx creado")


def create_inventario():
    """Create inventory template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Headers
    headers = ["Código Producto", "Nº Lote", "Cantidad", "Unidad", 
               "Fecha Fabricación", "Fecha Caducidad", "Ubicación"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Instructions row
    instructions = [
        "Debe existir",
        "Único por producto",
        "Numérico",
        "kg/l/ud",
        "AAAA-MM-DD",
        "AAAA-MM-DD (opcional)",
        "Código ubicación"
    ]
    for col, inst in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col, value=inst)
        cell.font = Font(italic=True, color="808080")
    
    # Example data
    examples = [
        ("MP-001", "AGUA-2025-001", 500, "l", "2025-01-01", "2027-01-01", "LIB"),
        ("MP-002", "GLI-2025-001", 80, "kg", "2025-01-01", "2026-06-01", "LIB"),
        ("ENV-001", "TARRO-2025-001", 1200, "ud", "2025-01-01", None, "LIB"),
        ("PA-001", "CR-2025-001", 120, "ud", "2025-01-01", "2026-01-01", "LIB"),
    ]
    
    for row, data in enumerate(examples, 3):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    style_header(ws, len(headers))
    auto_width(ws)
    
    wb.save("plantillas_importacion/inventario.xlsx")
    print("  ✓ inventario.xlsx creado")


def main():
    """Create all templates"""
    os.makedirs("plantillas_importacion", exist_ok=True)
    
    print("Generando plantillas Excel...")
    print()
    
    create_ubicaciones()
    create_productos()
    create_clientes()
    create_inventario()
    
    print()
    print("=" * 50)
    print("PLANTILLAS CREADAS EN: plantillas_importacion/")
    print("=" * 50)
    print()
    print("Instrucciones:")
    print("1. Edita los archivos Excel con los datos del cliente")
    print("2. Elimina las filas de ejemplo")
    print("3. Ejecuta: python import_data.py")
    print()


if __name__ == '__main__':
    main()
